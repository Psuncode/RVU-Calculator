#!/usr/bin/env python3
"""
Excel to JSON Parser for CMS RVU and GPCI Data
==============================================
Parses actual CMS Excel files (PPRRVU and GPCI) into JSON format for the RVU Calculator.

Usage:
    python parse_cms_excel.py PPRRVU22_JAN.xlsx data/rvu_data.json
    python parse_cms_excel.py GPCI2022.xlsx data/gpci_data.json
"""

import openpyxl
import json
import sys
from pathlib import Path


def parse_rvu_excel(excel_path, output_path, limit=None):
    """Parse CMS PPRRVU Excel file into JSON format"""

    print(f"Reading RVU Excel file: {excel_path}")
    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)

    # Get the first (and usually only) sheet
    sheet_name = wb.sheetnames[0]
    sheet = wb[sheet_name]
    print(f"Using sheet: {sheet_name}")

    # Find header row (contains "HCPCS" in first column)
    header_row = None
    for i, row in enumerate(sheet.iter_rows(min_row=1, max_row=20, values_only=True), 1):
        if row[0] and str(row[0]).strip().upper() == 'HCPCS':
            header_row = i
            print(f"Found header row at line {i}")
            print(f"Headers: {[str(cell).strip() if cell else '' for cell in row[:15]]}")
            break

    if not header_row:
        print("ERROR: Could not find header row with 'HCPCS'")
        sys.exit(1)

    # Column indices (0-based after we account for header row)
    # Based on PPRRVU22_JAN.xlsx inspection:
    # Col A=0: HCPCS
    # Col C=2: DESCRIPTION
    # Col F=5: WORK RVU
    # Col G=6: NON-FAC PE RVU
    # Col I=8: FACILITY PE RVU
    # Col K=10: MP RVU

    HCPCS_COL = 0
    DESC_COL = 2
    WORK_COL = 5
    PE_NONFAC_COL = 6
    PE_FAC_COL = 8
    MP_COL = 10

    rvu_data = {}
    skipped = 0
    parsed = 0

    print(f"\nParsing data rows (starting from row {header_row + 1})...")

    for row_num, row in enumerate(sheet.iter_rows(min_row=header_row + 1, values_only=True), header_row + 1):
        try:
            if not row or len(row) <= MP_COL:
                continue

            # Get CPT/HCPCS code
            cpt_code = row[HCPCS_COL]
            if not cpt_code:
                continue

            cpt_code = str(cpt_code).strip()

            # Skip non-alphanumeric codes or very long codes
            if not cpt_code or len(cpt_code) > 10:
                continue

            # Get description
            desc = str(row[DESC_COL]).strip() if row[DESC_COL] else ""

            # Parse numeric values
            def safe_float(val):
                if val is None or val == '':
                    return 0.0
                try:
                    return float(val)
                except (ValueError, TypeError):
                    return 0.0

            work_rvu = safe_float(row[WORK_COL])
            pe_nonfac = safe_float(row[PE_NONFAC_COL])
            pe_fac = safe_float(row[PE_FAC_COL])
            mp_rvu = safe_float(row[MP_COL])

            rvu_data[cpt_code] = {
                "desc": desc,
                "work_rvu": work_rvu,
                "pe_rvu_fac": pe_fac,
                "pe_rvu_nonfac": pe_nonfac,
                "mp_rvu": mp_rvu
            }

            parsed += 1

            # Show progress every 1000 rows
            if parsed % 1000 == 0:
                print(f"  Parsed {parsed} codes...")

            # Limit for testing
            if limit and parsed >= limit:
                print(f"  Reached limit of {limit} codes")
                break

        except Exception as e:
            skipped += 1
            if skipped <= 5:
                print(f"Warning: Skipped row {row_num}: {e}")

    print(f"\n✓ Parsed {parsed} CPT/HCPCS codes")
    if skipped > 0:
        print(f"  Skipped {skipped} rows due to errors")

    # Write JSON
    print(f"\nWriting JSON to: {output_path}")
    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    with open(output_path, 'w') as f:
        json.dump(rvu_data, f, indent=2)

    print(f"✓ Success! Created {output_path}")

    # Show sample codes
    print(f"\nSample codes included:")
    sample_codes = ['99213', '99214', '99215', '99203', '99204']
    for code in sample_codes:
        if code in rvu_data:
            print(f"  {code}: {rvu_data[code]['desc'][:60]}...")
            print(f"    Work: {rvu_data[code]['work_rvu']}, PE NF: {rvu_data[code]['pe_rvu_nonfac']}, PE F: {rvu_data[code]['pe_rvu_fac']}, MP: {rvu_data[code]['mp_rvu']}")

    return rvu_data


def parse_gpci_excel(excel_path, output_path):
    """Parse CMS GPCI Excel file into JSON format"""

    print(f"Reading GPCI Excel file: {excel_path}")
    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)

    sheet_name = wb.sheetnames[0]
    sheet = wb[sheet_name]
    print(f"Using sheet: {sheet_name}")

    # Find header row (contains "State" or "Locality Number")
    header_row = None
    for i, row in enumerate(sheet.iter_rows(min_row=1, max_row=10, values_only=True), 1):
        if row and len(row) > 2:
            # Check if this looks like the header row
            if any(cell and 'State' in str(cell) for cell in row):
                header_row = i
                print(f"Found header row at line {i}")
                print(f"Headers: {[str(cell).strip() if cell else '' for cell in row[:7]]}")
                break

    if not header_row:
        print("ERROR: Could not find header row")
        sys.exit(1)

    # Column indices based on GPCI2022.xlsx:
    # Col B=1: State
    # Col C=2: Locality Number
    # Col D=3: Locality Name
    # Col E=4: PW GPCI (Work)
    # Col F=5: PE GPCI
    # Col G=6: MP GPCI

    STATE_COL = 1
    LOCALITY_COL = 2
    NAME_COL = 3
    WORK_GPCI_COL = 4
    PE_GPCI_COL = 5
    MP_GPCI_COL = 6

    gpci_data = {}
    skipped = 0

    print(f"\nParsing GPCI data...")

    for row_num, row in enumerate(sheet.iter_rows(min_row=header_row + 1, values_only=True), header_row + 1):
        try:
            if not row or len(row) <= MP_GPCI_COL:
                continue

            state = str(row[STATE_COL]).strip() if row[STATE_COL] else ""
            locality = str(row[LOCALITY_COL]).strip() if row[LOCALITY_COL] else ""
            name = str(row[NAME_COL]).strip() if row[NAME_COL] else ""

            if not state or not locality:
                continue

            # Parse GPCI values
            def safe_float(val):
                if val is None or val == '':
                    return 1.000
                try:
                    return float(val)
                except (ValueError, TypeError):
                    return 1.000

            work_gpci = safe_float(row[WORK_GPCI_COL])
            pe_gpci = safe_float(row[PE_GPCI_COL])
            mp_gpci = safe_float(row[MP_GPCI_COL])

            # Create locality key (use state code for single-locality states, or state+locality for others)
            # For simplicity, we'll use just the state code as key
            locality_key = state

            # If this state already exists, we'll keep the first occurrence
            # (or you could create state-locality combinations)
            if locality_key not in gpci_data:
                gpci_data[locality_key] = {
                    "name": name,
                    "state": state,
                    "locality": locality,
                    "work_gpci": work_gpci,
                    "pe_gpci": pe_gpci,
                    "mp_gpci": mp_gpci
                }

        except Exception as e:
            skipped += 1
            if skipped <= 5:
                print(f"Warning: Skipped row {row_num}: {e}")

    print(f"\n✓ Parsed {len(gpci_data)} localities")
    if skipped > 0:
        print(f"  Skipped {skipped} rows")

    # Write JSON
    print(f"\nWriting JSON to: {output_path}")
    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    with open(output_path, 'w') as f:
        json.dump(gpci_data, f, indent=2)

    print(f"✓ Success! Created {output_path}")

    # Show Utah if present
    if 'UT' in gpci_data:
        print(f"\nUtah GPCI values:")
        print(f"  Locality: {gpci_data['UT']['locality']} - {gpci_data['UT']['name']}")
        print(f"  Work GPCI: {gpci_data['UT']['work_gpci']}")
        print(f"  PE GPCI:   {gpci_data['UT']['pe_gpci']}")
        print(f"  MP GPCI:   {gpci_data['UT']['mp_gpci']}")
    else:
        print("\nWARNING: Utah (UT) not found in GPCI data!")

    return gpci_data


def main():
    if len(sys.argv) < 3:
        print(__doc__)
        print("\nExamples:")
        print("  python parse_cms_excel.py PPRRVU22_JAN.xlsx data/rvu_data.json")
        print("  python parse_cms_excel.py GPCI2022.xlsx data/gpci_data.json")
        sys.exit(1)

    input_file = sys.argv[1]
    output_file = sys.argv[2]

    if not Path(input_file).exists():
        print(f"ERROR: Input file not found: {input_file}")
        sys.exit(1)

    # Detect file type
    input_lower = input_file.lower()

    if 'gpci' in input_lower:
        parse_gpci_excel(input_file, output_file)
    elif 'rvu' in input_lower or 'ppr' in input_lower:
        # Optional: Add limit parameter for testing
        limit = int(sys.argv[3]) if len(sys.argv) > 3 else None
        parse_rvu_excel(input_file, output_file, limit=limit)
    else:
        print("ERROR: Could not determine file type from filename.")
        print("Filename should contain 'GPCI' or 'RVU'/'PPR'")
        sys.exit(1)


if __name__ == '__main__':
    main()
