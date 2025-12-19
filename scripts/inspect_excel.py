#!/usr/bin/env python3
"""
Quick inspector to see the structure of CMS Excel files
"""
import openpyxl
import sys

def inspect_excel(filepath):
    print(f"\n{'='*60}")
    print(f"Inspecting: {filepath}")
    print('='*60)

    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)

    print(f"\nSheets: {wb.sheetnames}")

    for sheet_name in wb.sheetnames:
        print(f"\n--- Sheet: {sheet_name} ---")
        sheet = wb[sheet_name]

        # Get dimensions
        print(f"Max row: {sheet.max_row}, Max col: {sheet.max_column}")

        # Show first 10 rows
        print("\nFirst 10 rows:")
        for i, row in enumerate(sheet.iter_rows(values_only=True), 1):
            if i > 10:
                break
            # Clean up None values for display
            cleaned = [str(cell) if cell is not None else '' for cell in row[:15]]  # First 15 cols
            print(f"Row {i}: {cleaned}")

        print()

if __name__ == '__main__':
    if len(sys.argv) < 2:
        print("Usage: python inspect_excel.py <file.xlsx>")
        sys.exit(1)

    inspect_excel(sys.argv[1])
